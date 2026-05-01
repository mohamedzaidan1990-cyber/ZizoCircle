"""Build ZizoCircle 3-Year Financial Projections workbook."""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
import os

# ---- Colour palette ----
PURPLE      = "7B5CF6"
PURPLE_DARK = "5B3CD6"
PURPLE_LITE = "EDE7FE"
NAVY        = "0D0B1E"
WHITE       = "FFFFFF"
GREEN       = "16A34A"
RED         = "DC2626"
GREY_DARK   = "1A1A2E"
GREY_MID    = "6B7280"
GREY_LITE   = "F3F4F6"
LIME        = "C8F53B"
ORANGE      = "F5A53B"
PINK        = "F53B8F"

# ---- Number formats ----
FMT_QAR        = '"QAR " #,##0'
FMT_QAR_SIGNED = '"QAR " #,##0;[Red]"QAR " (#,##0)'
FMT_INT        = '#,##0'
FMT_PCT        = '0.0%'
FMT_PCT0       = '0%'
FMT_MULT       = '#,##0.0"x"'

# ---- Border helper ----
thin = Side(border_style='thin', color='E5E7EB')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---- Helpers ----
def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def apply_border_range(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = BORDER

def style_cell(cell, *, font=None, fill_color=None, alignment=None, number_format=None, border=True):
    if font is not None:
        cell.font = font
    if fill_color is not None:
        cell.fill = fill(fill_color)
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format
    if border:
        cell.border = BORDER


# ============================================================
# Workbook
# ============================================================
wb = Workbook()
wb.properties.title   = "Zizo Circle Financial Projections"
wb.properties.creator = "Bayline Holding W.L.L."
wb.properties.subject = "Snoonu Incubator Application"

# ============================================================
# TAB 1 — Executive Summary
# ============================================================
ws = wb.active
ws.title = "Executive Summary"
ws.sheet_properties.tabColor = PURPLE

# Default column widths
widths_t1 = {'A': 22, 'B': 16, 'C': 16, 'D': 22, 'E': 16, 'F': 16}
for col, w in widths_t1.items():
    ws.column_dimensions[col].width = w

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
right  = Alignment(horizontal='right',  vertical='center', wrap_text=True)

# Row 1 — Masthead
ws.merge_cells('A1:F1')
c = ws['A1']
c.value = "ZIZO CIRCLE — 3-YEAR FINANCIAL PROJECTIONS"
c.font = Font(name='Calibri', size=18, bold=True, color=WHITE)
c.fill = fill(NAVY)
c.alignment = center
ws.row_dimensions[1].height = 50

# Row 2 — Subtitle
ws.merge_cells('A2:F2')
c = ws['A2']
c.value = "Bayline Holding W.L.L. · Snoonu Incubator Application · May 2026"
c.font = Font(name='Calibri', size=10, italic=True, color=WHITE)
c.fill = fill(PURPLE)
c.alignment = center
ws.row_dimensions[2].height = 22

# Row 4 — MVP Live banner
ws.merge_cells('A4:F4')
c = ws['A4']
c.value = "● MVP LIVE  ·  Android APK Available  ·  Real Users Registered  ·  Firebase Backend Active"
c.font = Font(name='Calibri', size=12, bold=True, color=GREEN)
c.fill = fill(PURPLE_LITE)
c.alignment = center
ws.row_dimensions[4].height = 24

# Row 7 — FUNDING ASK heading
ws.merge_cells('A7:F7')
c = ws['A7']
c.value = "FUNDING ASK"
c.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
c.fill = fill(PURPLE)
c.alignment = center
ws.row_dimensions[7].height = 22

# Rows 8–11 — funding facts (label A:C, value D:F)
funding_rows = [
    ("Seed Funding Required", 1255000,        FMT_QAR),
    ("Pre-Money Valuation",   4000000,        FMT_QAR),
    ("Post-Money Valuation",  "=D8+D9",       FMT_QAR),
    ("Investor Equity",       "=D8/D10",      FMT_PCT),
]
for i, (label, value, nfmt) in enumerate(funding_rows):
    r = 8 + i
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    lc = ws.cell(row=r, column=1, value=label)
    lc.font = Font(name='Calibri', size=11, bold=True, color=GREY_DARK)
    lc.fill = fill(GREY_LITE)
    lc.alignment = left
    vc = ws.cell(row=r, column=4, value=value)
    vc.font = Font(name='Calibri', size=14, bold=True, color=PURPLE_DARK)
    vc.fill = fill(GREY_LITE)
    vc.alignment = right
    vc.number_format = nfmt
    ws.row_dimensions[r].height = 26
    # borders
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=col).fill = fill(GREY_LITE)

# Row 13 — Revenue outlook heading
ws.merge_cells('A13:F13')
c = ws['A13']
c.value = "3-YEAR REVENUE OUTLOOK (MEDIAN SCENARIO)"
c.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
c.fill = fill(PURPLE)
c.alignment = center
ws.row_dimensions[13].height = 22

# Rows 14–17 (header + 3 data rows) — table A:Year (merge A:B), C:Users, D:Revenue, E:F YoY (merge)
# Header row 14
header_t1 = [("A14:B14", "Year"), ("C14", "Users"), ("D14", "Revenue (QAR)"), ("E14:F14", "YoY Growth")]
ws.merge_cells('A14:B14')
ws.merge_cells('E14:F14')
ws['A14'] = "Year"
ws['C14'] = "Users"
ws['D14'] = "Revenue (QAR)"
ws['E14'] = "YoY Growth"
for col_letter in ('A','B','C','D','E','F'):
    cc = ws[f'{col_letter}14']
    cc.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    cc.fill = fill(PURPLE)
    cc.alignment = center
    cc.border = BORDER
ws.row_dimensions[14].height = 22

# Data rows 15,16,17  (alternating white / GREY_LITE)
data_t1 = [
    ("Year 1 — Qatar",         70000,   23500000,  None),
    ("Year 2 — UAE Expansion", 280000,  134000000, "=D16/D15-1"),
    ("Year 3 — KSA + Gulf",    900000,  562000000, "=D17/D16-1"),
]
for i, (yr, users, rev, yoy) in enumerate(data_t1):
    r = 15 + i
    row_fill = WHITE if i % 2 == 0 else GREY_LITE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value=yr)
    ws.cell(row=r, column=3, value=users).number_format = FMT_INT
    ws.cell(row=r, column=4, value=rev).number_format = FMT_QAR
    if yoy is None:
        ws.cell(row=r, column=5, value="—")
    else:
        ws.cell(row=r, column=5, value=yoy).number_format = FMT_PCT
    for col in range(1, 7):
        cc = ws.cell(row=r, column=col)
        cc.font = Font(name='Calibri', size=11, bold=(col == 1), color=GREY_DARK)
        cc.fill = fill(row_fill)
        cc.alignment = center if col != 1 else left
        cc.border = BORDER
    ws.row_dimensions[r].height = 22

# Row 18 — total spacer (left blank-ish)
# Row 20 — INVESTOR RETURNS heading
ws.merge_cells('A20:F20')
c = ws['A20']
c.value = "INVESTOR RETURNS"
c.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
c.fill = fill(PURPLE)
c.alignment = center
ws.row_dimensions[20].height = 22

# Rows 21–25 — big numbers
# Row 21: Investment
ws.merge_cells('A21:C21'); ws.merge_cells('D21:F21')
ws['A21'] = "Investment"
ws['A21'].font = Font(name='Calibri', size=11, bold=True, color=GREY_MID)
ws['A21'].fill = fill(WHITE); ws['A21'].alignment = left
ws['D21'] = "QAR 1,255,000"
ws['D21'].font = Font(name='Calibri', size=22, bold=True, color=PURPLE_DARK)
ws['D21'].fill = fill(WHITE); ws['D21'].alignment = right
ws.row_dimensions[21].height = 36

# Row 22: Implied Exit Valuation (basis row)
ws.merge_cells('A22:C22'); ws.merge_cells('D22:F22')
ws['A22'] = "Implied Exit Valuation"
ws['A22'].font = Font(name='Calibri', size=11, bold=True, color=GREY_MID)
ws['A22'].fill = fill(GREY_LITE); ws['A22'].alignment = left
ws['D22'] = 60912134
ws['D22'].number_format = FMT_QAR
ws['D22'].font = Font(name='Calibri', size=22, bold=True, color=PURPLE_DARK)
ws['D22'].fill = fill(GREY_LITE); ws['D22'].alignment = right
ws.row_dimensions[22].height = 36

# Row 23: 3-Year Exit Value (23.9% stake) = D22 * 0.239
ws.merge_cells('A23:C23'); ws.merge_cells('D23:F23')
ws['A23'] = "3-Year Exit Value (23.9% stake)"
ws['A23'].font = Font(name='Calibri', size=11, bold=True, color=GREY_MID)
ws['A23'].fill = fill(WHITE); ws['A23'].alignment = left
ws['D23'] = "=D22*0.239"
ws['D23'].number_format = FMT_QAR
ws['D23'].font = Font(name='Calibri', size=22, bold=True, color=PURPLE_DARK)
ws['D23'].fill = fill(WHITE); ws['D23'].alignment = right
ws.row_dimensions[23].height = 36

# Row 24: Multiple on Investment
ws.merge_cells('A24:C24'); ws.merge_cells('D24:F24')
ws['A24'] = "Multiple on Investment"
ws['A24'].font = Font(name='Calibri', size=11, bold=True, color=GREY_MID)
ws['A24'].fill = fill(GREY_LITE); ws['A24'].alignment = left
ws['D24'] = "=D23/1255000"
ws['D24'].number_format = FMT_MULT
ws['D24'].font = Font(name='Calibri', size=22, bold=True, color=PURPLE_DARK)
ws['D24'].fill = fill(GREY_LITE); ws['D24'].alignment = right
ws.row_dimensions[24].height = 36

# Row 25 — footnote
ws.merge_cells('A25:F25')
ws['A25'] = "based on conservative valuation methodology (≈0.108× Y3 revenue multiple, modest exit assumption)"
ws['A25'].font = Font(name='Calibri', size=8, italic=True, color=GREY_MID)
ws['A25'].alignment = center
ws.row_dimensions[25].height = 16

# Apply border to the big number rows
for r in (21, 22, 23, 24):
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = BORDER

ws.freeze_panes = 'A4'

# ============================================================
# TAB 2 — Year 1 Monthly P&L
# ============================================================
ws2 = wb.create_sheet("Year 1 Monthly P&L")
ws2.sheet_properties.tabColor = PURPLE

# Column widths
ws2.column_dimensions['A'].width = 36
for col_idx in range(2, 15):
    ws2.column_dimensions[get_column_letter(col_idx)].width = 14
ws2.column_dimensions['O'].width = 10

# Row 1 — masthead
ws2.merge_cells('A1:O1')
c = ws2['A1']
c.value = "ZIZO CIRCLE — YEAR 1 MONTHLY P&L (QATAR LAUNCH)"
c.font = Font(name='Calibri', size=16, bold=True, color=WHITE)
c.fill = fill(NAVY)
c.alignment = center
ws2.row_dimensions[1].height = 42

# Row 3 — column headers
months = ["May-26","Jun-26","Jul-26","Aug-26","Sep-26","Oct-26",
          "Nov-26","Dec-26","Jan-27","Feb-27","Mar-27","Apr-27"]
ws2.cell(row=3, column=1, value="Item")
for i, m in enumerate(months):
    ws2.cell(row=3, column=2+i, value=m)
ws2.cell(row=3, column=14, value="Total Y1")
ws2.cell(row=3, column=15, value="% of Rev")

for col in range(1, 16):
    cc = ws2.cell(row=3, column=col)
    cc.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    cc.fill = fill(PURPLE)
    cc.alignment = center
    cc.border = BORDER
ws2.row_dimensions[3].height = 22

# Row 4 — GROSS REVENUE
revenues = [400000, 600000, 900000, 1200000, 1500000, 1800000,
            2100000, 2300000, 2500000, 2800000, 3100000, 4300000]
ws2.cell(row=4, column=1, value="GROSS REVENUE (QAR)")
for i, v in enumerate(revenues):
    ws2.cell(row=4, column=2+i, value=v)
ws2.cell(row=4, column=14, value="=SUM(B4:M4)")
ws2.cell(row=4, column=15, value="=N4/$N$4")
for col in range(1, 16):
    cc = ws2.cell(row=4, column=col)
    cc.font = Font(name='Calibri', size=11, bold=True, color=GREY_DARK)
    cc.fill = fill(PURPLE_LITE)
    cc.alignment = left if col == 1 else right
    cc.border = BORDER
    if 2 <= col <= 14:
        cc.number_format = FMT_QAR_SIGNED
    elif col == 15:
        cc.number_format = FMT_PCT
ws2.row_dimensions[4].height = 22

# Rows 5–7 — revenue breakdown
breakdown = [
    ("  Cashback Commission (10% of GMV)",    0.70),
    ("  Premium Subscriptions (QAR 50/mo)",   0.20),
    ("  Featured Venue Listings",             0.10),
]
for i, (label, frac) in enumerate(breakdown):
    r = 5 + i
    ws2.cell(row=r, column=1, value=label)
    for col_idx in range(2, 14):
        col_letter = get_column_letter(col_idx)
        ws2.cell(row=r, column=col_idx, value=f"={col_letter}4*{frac}")
    ws2.cell(row=r, column=14, value=f"=SUM(B{r}:M{r})")
    ws2.cell(row=r, column=15, value=f"=N{r}/$N$4")
    for col in range(1, 16):
        cc = ws2.cell(row=r, column=col)
        cc.font = Font(name='Calibri', size=10, italic=True, color=GREY_MID)
        cc.alignment = left if col == 1 else right
        cc.border = BORDER
        if 2 <= col <= 14:
            cc.number_format = FMT_QAR_SIGNED
        elif col == 15:
            cc.number_format = FMT_PCT

# Row 8 — blank
# Row 9 — OPERATING EXPENSES heading
ws2.cell(row=9, column=1, value="OPERATING EXPENSES")
for col in range(1, 16):
    cc = ws2.cell(row=9, column=col)
    cc.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    cc.fill = fill(GREY_DARK)
    cc.alignment = left if col == 1 else center
    cc.border = BORDER
ws2.row_dimensions[9].height = 22

# Rows 10–13 — OPEX line items
opex_lines = [
    ("  Team Salaries",                  40000),
    ("  Office & Utilities",             15000),
    ("  Marketing & User Acquisition",   41700),
    ("  Operations / Legal / Misc",      2500),
]
for i, (label, monthly) in enumerate(opex_lines):
    r = 10 + i
    ws2.cell(row=r, column=1, value=label)
    for col_idx in range(2, 14):
        ws2.cell(row=r, column=col_idx, value=monthly)
    ws2.cell(row=r, column=14, value=f"=SUM(B{r}:M{r})")
    ws2.cell(row=r, column=15, value=f"=N{r}/$N$4")
    for col in range(1, 16):
        cc = ws2.cell(row=r, column=col)
        cc.font = Font(name='Calibri', size=11, color=GREY_DARK)
        cc.alignment = left if col == 1 else right
        cc.border = BORDER
        if 2 <= col <= 14:
            cc.number_format = FMT_QAR_SIGNED
        elif col == 15:
            cc.number_format = FMT_PCT
        if i % 2 == 1:
            cc.fill = fill(GREY_LITE)

# Row 14 — TOTAL OPEX
ws2.cell(row=14, column=1, value="TOTAL OPEX")
for col_idx in range(2, 14):
    col_letter = get_column_letter(col_idx)
    ws2.cell(row=14, column=col_idx, value=f"=SUM({col_letter}10:{col_letter}13)")
ws2.cell(row=14, column=14, value="=SUM(N10:N13)")
ws2.cell(row=14, column=15, value="=N14/$N$4")
for col in range(1, 16):
    cc = ws2.cell(row=14, column=col)
    cc.font = Font(name='Calibri', size=11, bold=True, color=RED)
    cc.fill = fill(PURPLE_LITE)
    cc.alignment = left if col == 1 else right
    cc.border = BORDER
    if 2 <= col <= 14:
        cc.number_format = FMT_QAR_SIGNED
    elif col == 15:
        cc.number_format = FMT_PCT
ws2.row_dimensions[14].height = 22

# Row 16 — GROSS PROFIT
ws2.cell(row=16, column=1, value="GROSS PROFIT (Revenue − OPEX)")
for col_idx in range(2, 14):
    col_letter = get_column_letter(col_idx)
    ws2.cell(row=16, column=col_idx, value=f"={col_letter}4-{col_letter}14")
ws2.cell(row=16, column=14, value="=N4-N14")
ws2.cell(row=16, column=15, value="=N16/$N$4")
for col in range(1, 16):
    cc = ws2.cell(row=16, column=col)
    cc.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    cc.fill = fill(PURPLE)
    cc.alignment = left if col == 1 else right
    cc.border = BORDER
    if 2 <= col <= 14:
        cc.number_format = FMT_QAR_SIGNED
    elif col == 15:
        cc.number_format = FMT_PCT
ws2.row_dimensions[16].height = 22

# Row 17 — Profit Margin %
ws2.cell(row=17, column=1, value="PROFIT MARGIN %")
for col_idx in range(2, 14):
    col_letter = get_column_letter(col_idx)
    ws2.cell(row=17, column=col_idx, value=f"={col_letter}16/{col_letter}4")
ws2.cell(row=17, column=14, value="=N16/N4")
ws2.cell(row=17, column=15, value="")
for col in range(1, 16):
    cc = ws2.cell(row=17, column=col)
    cc.font = Font(name='Calibri', size=11, bold=True, color=GREY_DARK)
    cc.fill = fill(GREY_LITE)
    cc.alignment = left if col == 1 else right
    cc.border = BORDER
    if 2 <= col <= 14:
        cc.number_format = FMT_PCT
ws2.row_dimensions[17].height = 22

ws2.freeze_panes = 'B4'

# ============================================================
# TAB 3 — 3-Year Projections
# ============================================================
ws3 = wb.create_sheet("3-Year Projections")
ws3.sheet_properties.tabColor = PURPLE

# Column widths
ws3.column_dimensions['A'].width = 22
for col_idx in range(2, 11):  # B–J
    ws3.column_dimensions[get_column_letter(col_idx)].width = 15
ws3.column_dimensions['K'].width = 16

# Row 1 — masthead
ws3.merge_cells('A1:K1')
c = ws3['A1']
c.value = "ZIZO CIRCLE — 3-YEAR PROJECTIONS BY SCENARIO"
c.font = Font(name='Calibri', size=16, bold=True, color=WHITE)
c.fill = fill(NAVY)
c.alignment = center
ws3.row_dimensions[1].height = 42

# Row 3 — Scenario header banners
ws3.merge_cells('B3:D3'); ws3['B3'] = "CONSERVATIVE"
ws3.merge_cells('E3:G3'); ws3['E3'] = "MEDIAN"
ws3.merge_cells('H3:J3'); ws3['H3'] = "HIGH"
ws3['K3'] = "3-Yr Total"

for cell_addr, color in [('B3', ORANGE), ('E3', PURPLE), ('H3', PINK), ('K3', NAVY)]:
    cc = ws3[cell_addr]
    cc.font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    cc.fill = fill(color)
    cc.alignment = center
# borders for full row
for col in range(1, 12):
    ws3.cell(row=3, column=col).border = BORDER
ws3.row_dimensions[3].height = 24

# Row 4 — sub-headers
ws3['A4'] = "Metric"
sub_headers = ["Year 1", "Year 2", "Year 3"] * 3
for i, h in enumerate(sub_headers):
    ws3.cell(row=4, column=2+i, value=h)
ws3.cell(row=4, column=11, value="")  # K4 spacer

for col in range(1, 12):
    cc = ws3.cell(row=4, column=col)
    cc.font = Font(name='Calibri', size=11, bold=True, color=GREY_DARK)
    cc.fill = fill(PURPLE_LITE)
    cc.alignment = center
    cc.border = BORDER
ws3.row_dimensions[4].height = 22

# Row 5 — Users (cumulative)
users_data = {
    'B5': 40000,  'C5': 150000, 'D5': 500000,    # Conservative
    'E5': 70000,  'F5': 280000, 'G5': 900000,    # Median
    'H5': 100000, 'I5': 450000, 'J5': 1500000,   # High
}
ws3['A5'] = "Users (cumulative)"
for k, v in users_data.items():
    ws3[k] = v
    ws3[k].number_format = FMT_INT
ws3['K5'] = "=G5"  # end-of-Y3 cumulative for median
ws3['K5'].number_format = FMT_INT

# Row 6 — Revenue (QAR)
rev_data = {
    'B6': 9600000,   'C6': 54000000,  'D6': 240000000,
    'E6': 23500000,  'F6': 134000000, 'G6': 562000000,
    'H6': 42000000,  'I6': 270000000, 'J6': 1170000000,
}
ws3['A6'] = "Revenue (QAR)"
for k, v in rev_data.items():
    ws3[k] = v
    ws3[k].number_format = FMT_QAR
ws3['K6'] = "=E6+F6+G6"  # median 3-yr total
ws3['K6'].number_format = FMT_QAR

# Row 7 — Revenue/User
ws3['A7'] = "Revenue/User"
for col_idx in range(2, 11):
    col_letter = get_column_letter(col_idx)
    ws3.cell(row=7, column=col_idx, value=f"={col_letter}6/{col_letter}5")
    ws3.cell(row=7, column=col_idx).number_format = FMT_QAR

# Row 8 — YoY Growth %
ws3['A8'] = "YoY Growth %"
for grp_start in (2, 5, 8):  # B, E, H
    y1 = get_column_letter(grp_start)
    y2 = get_column_letter(grp_start + 1)
    y3 = get_column_letter(grp_start + 2)
    ws3.cell(row=8, column=grp_start, value="—")
    ws3.cell(row=8, column=grp_start + 1, value=f"={y2}6/{y1}6-1")
    ws3.cell(row=8, column=grp_start + 2, value=f"={y3}6/{y2}6-1")
    ws3.cell(row=8, column=grp_start + 1).number_format = FMT_PCT0
    ws3.cell(row=8, column=grp_start + 2).number_format = FMT_PCT0

# Row 9 — OPEX
opex_data = {
    'B9': 1190400, 'C9': 4500000,  'D9': 18000000,
    'E9': 1190400, 'F9': 6000000,  'G9': 28000000,
    'H9': 1190400, 'I9': 8500000,  'J9': 45000000,
}
ws3['A9'] = "OPEX (QAR)"
for k, v in opex_data.items():
    ws3[k] = v
    ws3[k].number_format = FMT_QAR

# Row 10 — EBITDA = Revenue − OPEX
ws3['A10'] = "EBITDA (QAR)"
for col_idx in range(2, 11):
    col_letter = get_column_letter(col_idx)
    ws3.cell(row=10, column=col_idx, value=f"={col_letter}6-{col_letter}9")
    ws3.cell(row=10, column=col_idx).number_format = FMT_QAR_SIGNED

# Row 11 — EBITDA Margin %
ws3['A11'] = "EBITDA Margin %"
for col_idx in range(2, 11):
    col_letter = get_column_letter(col_idx)
    ws3.cell(row=11, column=col_idx, value=f"={col_letter}10/{col_letter}6")
    ws3.cell(row=11, column=col_idx).number_format = FMT_PCT0

# Style data rows 5–11
for r in range(5, 12):
    # Column A bold
    cc = ws3.cell(row=r, column=1)
    cc.font = Font(name='Calibri', size=11, bold=True, color=GREY_DARK)
    cc.alignment = left
    cc.fill = fill(WHITE if r % 2 == 1 else GREY_LITE)
    cc.border = BORDER
    # Data columns
    for col_idx in range(2, 12):
        cc = ws3.cell(row=r, column=col_idx)
        cc.font = Font(name='Calibri', size=11, color=GREY_DARK)
        cc.alignment = right
        cc.border = BORDER
        # Tint median band slightly
        if 5 <= col_idx <= 7:
            cc.fill = fill(PURPLE_LITE if r % 2 == 0 else WHITE)
        else:
            cc.fill = fill(GREY_LITE if r % 2 == 0 else WHITE)
    ws3.row_dimensions[r].height = 22

# Row 13 — Market milestones banner
ws3.merge_cells('A13:K13')
c = ws3['A13']
c.value = "MARKET MILESTONES"
c.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
c.fill = fill(PURPLE)
c.alignment = center
ws3.row_dimensions[13].height = 22
for col in range(1, 12):
    ws3.cell(row=13, column=col).border = BORDER

# Row 14 — Markets
ws3['A14'] = "Market"
ws3.merge_cells('B14:D14'); ws3['B14'] = "🇶🇦 Qatar"
ws3.merge_cells('E14:G14'); ws3['E14'] = "🇶🇦 Qatar → 🇦🇪 UAE"
ws3.merge_cells('H14:J14'); ws3['H14'] = "🇶🇦🇦🇪 → 🇸🇦 KSA + Gulf"
ws3['K14'] = ""
for cell_addr in ('A14','B14','E14','H14','K14'):
    cc = ws3[cell_addr]
    cc.font = Font(name='Calibri', size=11, bold=True, color=GREY_DARK)
    cc.fill = fill(PURPLE_LITE)
    cc.alignment = center
for col in range(1, 12):
    ws3.cell(row=14, column=col).border = BORDER
    if ws3.cell(row=14, column=col).fill.start_color.rgb in (None, '00000000'):
        ws3.cell(row=14, column=col).fill = fill(PURPLE_LITE)
ws3.row_dimensions[14].height = 26

ws3.freeze_panes = 'B5'

# ============================================================
# TAB 4 — Use of Funds
# ============================================================
ws4 = wb.create_sheet("Use of Funds")
ws4.sheet_properties.tabColor = PURPLE

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 44
ws4.column_dimensions['E'].width = 2
for col_letter in ('F','G','H','I','J','K'):
    ws4.column_dimensions[col_letter].width = 12

# Row 1 — masthead
ws4.merge_cells('A1:F1')
c = ws4['A1']
c.value = "USE OF FUNDS — QAR 1,255,000 SEED"
c.font = Font(name='Calibri', size=16, bold=True, color=WHITE)
c.fill = fill(NAVY)
c.alignment = center
ws4.row_dimensions[1].height = 42

# Row 3 — header
headers_t4 = ["Category", "Amount (QAR)", "% of Total", "Notes / Use"]
for i, h in enumerate(headers_t4):
    cc = ws4.cell(row=3, column=1+i, value=h)
    cc.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    cc.fill = fill(PURPLE)
    cc.alignment = center
    cc.border = BORDER
ws4.row_dimensions[3].height = 24

# Rows 4–9 — data
funds = [
    ("Marketing & User Acquisition", 500000, "Influencer campaigns, performance ads, brand"),
    ("Team Salaries (12 months)",    480000, "Founder + 4 team members"),
    ("Office & Operations",          180000, "Lusail co-working space, utilities"),
    ("iOS Build + App Store",         50000, "Apple Developer, code signing, polish"),
    ("Legal & Compliance",            30000, "MOCI registration, contracts"),
    ("Contingency / Buffer",          15000, "Reserve for unexpected costs"),
]
for i, (cat, amt, notes) in enumerate(funds):
    r = 4 + i
    ws4.cell(row=r, column=1, value=cat)
    ws4.cell(row=r, column=2, value=amt).number_format = FMT_QAR
    ws4.cell(row=r, column=3, value=f"=B{r}/$B$10").number_format = FMT_PCT
    ws4.cell(row=r, column=4, value=notes)
    row_fill = WHITE if i % 2 == 0 else GREY_LITE
    for col in range(1, 5):
        cc = ws4.cell(row=r, column=col)
        cc.font = Font(name='Calibri', size=11, color=GREY_DARK)
        cc.fill = fill(row_fill)
        cc.alignment = left if col in (1, 4) else right
        cc.border = BORDER
    ws4.row_dimensions[r].height = 22

# Row 10 — TOTAL
ws4.cell(row=10, column=1, value="TOTAL")
ws4.cell(row=10, column=2, value="=SUM(B4:B9)").number_format = FMT_QAR
ws4.cell(row=10, column=3, value=1.0).number_format = FMT_PCT
ws4.cell(row=10, column=4, value="")
for col in range(1, 5):
    cc = ws4.cell(row=10, column=col)
    cc.font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    cc.fill = fill(PURPLE)
    cc.alignment = left if col in (1, 4) else right
    cc.border = BORDER
ws4.row_dimensions[10].height = 26

# ---- Pie chart ----
pie = PieChart()
labels_ref = Reference(ws4, min_col=1, min_row=4, max_row=9)
data_ref   = Reference(ws4, min_col=2, min_row=3, max_row=9)
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(labels_ref)
pie.title = "Allocation of QAR 1.255M Seed"
pie.dataLabels = DataLabelList(showPercent=True)

# Brand-tinted slice colours
from openpyxl.chart.shapes import GraphicalProperties
slice_colors = [PURPLE, PURPLE_DARK, PINK, ORANGE, LIME, PURPLE_LITE]
series = pie.series[0]
data_points = []
for i, col in enumerate(slice_colors):
    dp = DataPoint(idx=i)
    dp.graphicalProperties = GraphicalProperties(solidFill=col)
    data_points.append(dp)
series.data_points = data_points

pie.height = 11
pie.width  = 16
ws4.add_chart(pie, "F3")

# ============================================================
# TAB 5 — Investor Returns
# ============================================================
ws5 = wb.create_sheet("Investor Returns")
ws5.sheet_properties.tabColor = PURPLE

ws5.column_dimensions['A'].width = 24
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 18
ws5.column_dimensions['D'].width = 20
ws5.column_dimensions['E'].width = 22
ws5.column_dimensions['F'].width = 22

# Row 1 — masthead
ws5.merge_cells('A1:E1')
c = ws5['A1']
c.value = "INVESTOR RETURNS ANALYSIS"
c.font = Font(name='Calibri', size=16, bold=True, color=WHITE)
c.fill = fill(NAVY)
c.alignment = center
ws5.row_dimensions[1].height = 42

# Row 3 — DEAL STRUCTURE banner
ws5.merge_cells('A3:E3')
c = ws5['A3']
c.value = "DEAL STRUCTURE"
c.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
c.fill = fill(PURPLE)
c.alignment = center
ws5.row_dimensions[3].height = 22
for col in range(1, 6):
    ws5.cell(row=3, column=col).border = BORDER

# Rows 4–8 — deal facts
deal_rows = [
    ("Investment Amount",     1255000,       FMT_QAR),
    ("Pre-Money Valuation",   4000000,       FMT_QAR),
    ("Post-Money Valuation",  "=B5+B4",      FMT_QAR),
    ("Equity Granted",        "=B4/B6",      FMT_PCT),
    ("Investment Date",       "May 2026",    None),
]
for i, (label, value, nfmt) in enumerate(deal_rows):
    r = 4 + i
    lc = ws5.cell(row=r, column=1, value=label)
    lc.font = Font(name='Calibri', size=12, bold=True, color=GREY_DARK)
    lc.alignment = left
    lc.fill = fill(GREY_LITE if i % 2 == 1 else WHITE)
    lc.border = BORDER
    vc = ws5.cell(row=r, column=2, value=value)
    vc.font = Font(name='Calibri', size=14, bold=True, color=PURPLE_DARK)
    vc.alignment = right
    vc.fill = fill(GREY_LITE if i % 2 == 1 else WHITE)
    vc.border = BORDER
    if nfmt:
        vc.number_format = nfmt
    # Padding cells C–E in same row
    for col in range(3, 6):
        cc = ws5.cell(row=r, column=col)
        cc.fill = fill(GREY_LITE if i % 2 == 1 else WHITE)
        cc.border = BORDER
    ws5.row_dimensions[r].height = 24

# Row 10 — Exit scenarios banner
ws5.merge_cells('A10:F10')
c = ws5['A10']
c.value = "EXIT SCENARIOS (3-YEAR HORIZON)"
c.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
c.fill = fill(PURPLE)
c.alignment = center
ws5.row_dimensions[10].height = 22
for col in range(1, 7):
    ws5.cell(row=10, column=col).border = BORDER

# Row 11 — exit table headers
exit_headers = ["Scenario", "Y3 Revenue", "Revenue Multiple", "Implied Valuation",
                "Investor 23.9% Stake", "Multiple on Investment"]
for i, h in enumerate(exit_headers):
    cc = ws5.cell(row=11, column=1+i, value=h)
    cc.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    cc.fill = fill(PURPLE_DARK)
    cc.alignment = center
    cc.border = BORDER
ws5.row_dimensions[11].height = 26

# Rows 12–14 — exit scenarios
exit_rows = [
    ("Conservative", 240000000,  0.10),
    ("Median",       562000000,  0.108),
    ("High",         1170000000, 0.10),
]
for i, (name, rev, mult) in enumerate(exit_rows):
    r = 12 + i
    ws5.cell(row=r, column=1, value=name)
    ws5.cell(row=r, column=2, value=rev).number_format = FMT_QAR
    ws5.cell(row=r, column=3, value=mult).number_format = FMT_MULT
    ws5.cell(row=r, column=4, value=f"=B{r}*C{r}").number_format = FMT_QAR
    ws5.cell(row=r, column=5, value=f"=D{r}*0.239").number_format = FMT_QAR
    ws5.cell(row=r, column=6, value=f"=E{r}/1255000").number_format = FMT_MULT

    is_median = (name == "Median")
    row_fill = PURPLE_LITE if is_median else (GREY_LITE if i % 2 == 1 else WHITE)
    for col in range(1, 7):
        cc = ws5.cell(row=r, column=col)
        if is_median and col == 6:
            cc.font = Font(name='Calibri', size=14, bold=True, color=PURPLE_DARK)
        elif is_median:
            cc.font = Font(name='Calibri', size=11, bold=True, color=PURPLE_DARK)
        else:
            cc.font = Font(name='Calibri', size=11, color=GREY_DARK)
        cc.fill = fill(row_fill)
        cc.alignment = left if col == 1 else right
        cc.border = BORDER
    ws5.row_dimensions[r].height = 24

# Row 16 — RETURN SUMMARY banner
ws5.merge_cells('A16:E16')
c = ws5['A16']
c.value = "RETURN SUMMARY (MEDIAN CASE)"
c.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
c.fill = fill(PURPLE)
c.alignment = center
ws5.row_dimensions[16].height = 22
for col in range(1, 6):
    ws5.cell(row=16, column=col).border = BORDER

# Rows 17–20 — big numbers
big_numbers = [
    ("Investment",              "QAR 1,255,000",      None,          False),
    ("3-Year Exit Value",       "QAR 14,558,000",     None,          False),
    ("Multiple of Investment",  "11.6×",              None,          False),
    ("IRR (approx)",            "=(14558000/1255000)^(1/3)-1", FMT_PCT, True),
]
for i, (label, value, nfmt, is_formula) in enumerate(big_numbers):
    r = 17 + i
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws5.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    lc = ws5.cell(row=r, column=1, value=label)
    lc.font = Font(name='Calibri', size=10, color=GREY_MID)
    lc.alignment = left
    lc.fill = fill(WHITE if i % 2 == 0 else GREY_LITE)
    vc = ws5.cell(row=r, column=3, value=value)
    vc.font = Font(name='Calibri', size=24, bold=True, color=PURPLE_DARK)
    vc.alignment = right
    vc.fill = fill(WHITE if i % 2 == 0 else GREY_LITE)
    if nfmt:
        vc.number_format = nfmt
    for col in range(1, 6):
        cc = ws5.cell(row=r, column=col)
        cc.border = BORDER
        cc.fill = fill(WHITE if i % 2 == 0 else GREY_LITE)
    ws5.row_dimensions[r].height = 38

# ============================================================
# Save
# ============================================================
out_path = r"C:\Users\User\Downloads\ZizoCircle_Financials.xlsx"
wb.save(out_path)

# ============================================================
# Verification & summary
# ============================================================
size_kb = os.path.getsize(out_path) / 1024
print(f"File saved: {out_path}")
print(f"File size:  {size_kb:.1f} KB")
print(f"Sheets:     {wb.sheetnames}")

# Summary dictionary of key totals
y1_revenue = sum([400000, 600000, 900000, 1200000, 1500000, 1800000,
                  2100000, 2300000, 2500000, 2800000, 3100000, 4300000])
y1_opex_total = (40000 + 15000 + 41700 + 2500) * 12
post_money = 1255000 + 4000000
equity_pct = 1255000 / post_money
return_multiple = 14558000 / 1255000

summary = {
    "Y1 revenue (QAR)":          f"{y1_revenue:,}",
    "Y3 revenue median (QAR)":   f"{562000000:,}",
    "Y1 total OPEX (QAR)":       f"{y1_opex_total:,}",
    "Equity %":                   f"{equity_pct*100:.1f}%",
    "Return multiple (median)":   f"{return_multiple:.1f}x",
}
print("\nKey totals for verification:")
for k, v in summary.items():
    print(f"  {k:30s}  {v}")
